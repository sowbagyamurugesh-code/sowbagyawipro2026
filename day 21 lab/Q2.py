import pandas as pd
# using pandas
df=pd.read_excel("sales_data.xlsx")
df["Total"]=df["Quantity"]*df["Price"]
print(df)
# Save updated DataFrame into new Excel file
df.to_excel("sales_summary.xlsx", index=False)
print("sales_summary.xlsx created successfully using pandas!")



from openpyxl import load_workbook, Workbook
# Load existing workbook
wb = load_workbook("sales_data.xlsx")
sheet = wb["Sheet1"]
# Create new workbook for output
new_wb = Workbook()
new_sheet = new_wb.active
new_sheet.title = "Sales Summary"

# Read header row
headers = []
for cell in sheet[1]:
    headers.append(cell.value)

# Add "Total" column
headers.append("Total")
new_sheet.append(headers)

# Read data rows and calculate Total
for row in sheet.iter_rows(min_row=2, values_only=True):
    product, quantity, price = row
    total = quantity * price
    new_sheet.append([product, quantity, price, total])

# Save output file
new_wb.save("Sales_Summary_openpyxl.xlsx")

print("Sales_Summary_openpyxl.xlsx created successfully using OpenPyXL!")
