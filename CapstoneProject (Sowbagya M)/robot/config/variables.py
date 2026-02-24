# config/variables.py
#
# Reads users_test_data.xlsx "Test Data" sheet and exposes:
#   USERS  — list of dicts, one per row, keys = column headers
#   BROWSER, APP_URL — runtime settings
#
# Robot Framework loads this file via:
#   Variables    ../config/variables.py
# and then ${USERS} is available in every test/keyword.

import os
import openpyxl

# ── Runtime settings ──────────────────────────────────────────────────────────
BROWSER           = "chrome"
APP_URL           = "https://automationexercise.com"
IMPLICIT_WAIT     = "0s"
PAGE_LOAD_TIMEOUT = "30s"

# ── Read Excel → USERS list ───────────────────────────────────────────────────
def _load_users():
    # Resolve path relative to this file so it works from any working directory
    here     = os.path.dirname(os.path.abspath(__file__))
    xlsx     = os.path.join(here, '..', 'test_data', 'users_test_data.xlsx')
    wb       = openpyxl.load_workbook(xlsx)
    ws       = wb['Test Data']

    headers  = [cell.value for cell in ws[1]]          # row 1 = column names
    users    = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):                               # skip blank rows
            continue
        # Build dict: {"name": "John Smith", "email": "...", ...}
        user = {headers[i]: str(row[i]) for i in range(len(headers))}
        users.append(user)
    return users

USERS = _load_users()
